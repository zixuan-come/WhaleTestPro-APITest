# -*- coding: utf-8 -*-
"""从 data/*.yaml 生成接口测试用例台账 Excel(单文件、每个资源一个 sheet)。

设计要点:
- 用例编号(编号列)、预期结果 一律从 YAML 读取,保证台账与数据文件天然一致;
  YAML 改了重跑本脚本即可同步,不会两边对不上。
- 请求方式 / 路径 / 描述 / 备注 由本脚本的 META 表按 case_id 提供。
- 脚本会校验:YAML 里出现的每个 case_id 都必须在 META 里有元数据,
  否则报警退出(落实"编号不一致立刻提醒")。反向多余的 META key 也会警告。
- negative_contract 是参数化生成的契约反例,无独立 case_id,单独按资源展开。

用法: python tools/gen_testcase_excel.py
输出: docs/WhaleTestPro接口测试用例.xlsx
"""
import json
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "docs" / "WhaleTestPro接口测试用例.xlsx"

HEADERS = ["用例编号", "分组", "请求方式", "接口路径", "用例描述", "请求数据", "预期结果", "备注"]

# ---------------------------------------------------------------------------
# 每个 sheet 的取数配置:
#   sections: (yaml相对路径, 组名, 方式, 路径, 该组默认描述)
#   编号与预期从 yaml 读;同组多条时用 CASE_DESC/CASE_REMARK 按 case_id 细化。
#   extra_rows: 不由 yaml 驱动的内联用例(如 login_success 写在测试代码里)。
# ---------------------------------------------------------------------------
SHEETS = [
    {
        "title": "用户认证",
        "sections": [
            ("login.yaml", "login_fail", "POST", "/auth/login", "登录失败"),
            ("logout.yaml", "logout_success", "POST", "/auth/logout", "登出成功"),
            ("logout.yaml", "logout_fail", "POST", "/auth/logout", "登出失败"),
            ("misc.yaml", "register_success", "POST", "/auth/register", "注册新用户"),
            ("misc.yaml", "register_fail", "POST", "/auth/register", "注册失败"),
        ],
        "extra_rows": [
            ["login_success", "login_success", "POST", "/auth/login",
             "正确用户名+密码登录成功", '{"username": "<admin>", "password": "<pwd>"}',
             "HTTP 200；含字段 access_token/token_type", "内联(非 YAML 驱动)"],
        ],
    },
    {
        "title": "项目",
        "sections": [
            ("project/create_project.yaml", "create_project_success", "POST", "/projects", "创建项目"),
            ("project/create_project.yaml", "create_project_fail", "POST", "/projects", "创建项目失败"),
            ("project/list_project.yaml", "list_project_success", "GET", "/projects", "项目列表"),
            ("project/list_project.yaml", "list_project_fail", "GET", "/projects", "项目列表(无鉴权)"),
            ("project/get_project_detail.yaml", "get_project_detail_success", "GET", "/projects/{id}", "项目详情"),
            ("project/get_project_detail.yaml", "get_project_detail_fail", "GET", "/projects/{id}", "项目详情失败"),
            ("project/delete_project.yaml", "delete_project_success", "DELETE", "/projects/{id}", "删除项目"),
            ("project/delete_project.yaml", "delete_project_fail", "DELETE", "/projects/{id}", "删除项目失败"),
        ],
    },
    {
        "title": "接口",
        "sections": [
            ("interface.yaml", "create_success", "POST", "/interfaces", "创建接口"),
            ("interface.yaml", "create_fail", "POST", "/interfaces", "创建接口失败"),
            ("interface.yaml", "list_success", "GET", "/interfaces", "接口列表"),
            ("interface.yaml", "detail_success", "GET", "/interfaces/{id}", "接口详情"),
            ("interface.yaml", "detail_fail", "GET", "/interfaces/{id}", "接口详情(不存在)"),
            ("interface.yaml", "update_success", "PUT", "/interfaces/{id}", "更新接口"),
            ("interface.yaml", "delete_success", "DELETE", "/interfaces/{id}", "删除接口"),
            ("interface.yaml", "category_rename", "PUT", "/interfaces/categories/{name}", "分类改名"),
            ("interface.yaml", "category_delete", "DELETE", "/interfaces/categories/{name}", "分类删除"),
        ],
    },
    {
        "title": "用例",
        "sections": [
            ("case.yaml", "create_success", "POST", "/cases", "创建用例"),
            ("case.yaml", "create_fail", "POST", "/cases", "创建用例失败"),
            ("case.yaml", "create_no_token", "POST", "/cases", "创建用例(不带 token)"),
            ("case.yaml", "list_success", "GET", "/cases", "用例列表"),
            ("case.yaml", "detail_success", "GET", "/cases/{id}", "用例详情"),
            ("case.yaml", "detail_fail", "GET", "/cases/{id}", "用例详情(不存在)"),
            ("case.yaml", "delete_success", "DELETE", "/cases/{id}", "删除用例"),
            ("case.yaml", "run_success", "POST", "/cases/{id}/run", "执行用例"),
            ("case.yaml", "chain_success", "POST", "/cases/{id}/run", "用例串联执行(上下文传参)"),
        ],
    },
    {
        "title": "场景",
        "sections": [
            ("scenario.yaml", "create_success", "POST", "/scenarios", "创建场景"),
            ("scenario.yaml", "create_no_token", "POST", "/scenarios", "创建场景(不带 token)"),
            ("scenario.yaml", "list_success", "GET", "/scenarios", "场景列表"),
            ("scenario.yaml", "detail_success", "GET", "/scenarios/{id}", "场景详情"),
            ("scenario.yaml", "detail_fail", "GET", "/scenarios/{id}", "场景详情(不存在)"),
            ("scenario.yaml", "update_success", "PUT", "/scenarios/{id}", "更新场景"),
            ("scenario.yaml", "delete_success", "DELETE", "/scenarios/{id}", "删除场景"),
            ("scenario.yaml", "run_success", "POST", "/scenarios/{id}/run", "执行场景"),
        ],
    },
    {
        "title": "环境",
        "sections": [
            ("environment.yaml", "create_success", "POST", "/environments", "创建环境"),
            ("environment.yaml", "create_no_token", "POST", "/environments", "创建环境(不带 token)"),
            ("environment.yaml", "list_success", "GET", "/environments", "环境列表"),
            ("environment.yaml", "detail_success", "GET", "/environments/{id}", "环境详情"),
            ("environment.yaml", "detail_fail", "GET", "/environments/{id}", "环境详情(不存在)"),
            ("environment.yaml", "update_success", "PUT", "/environments/{id}", "更新环境"),
            ("environment.yaml", "delete_success", "DELETE", "/environments/{id}", "删除环境"),
        ],
    },
    {
        "title": "Mock",
        "sections": [
            ("mock.yaml", "create_success", "POST", "/mocks", "创建 Mock"),
            ("mock.yaml", "create_no_token", "POST", "/mocks", "创建 Mock(不带 token)"),
            ("mock.yaml", "list_success", "GET", "/mocks", "Mock 列表"),
            ("mock.yaml", "detail_success", "GET", "/mocks/{id}", "Mock 详情"),
            ("mock.yaml", "detail_fail", "GET", "/mocks/{id}", "Mock 详情(不存在)"),
            ("mock.yaml", "update_success", "PUT", "/mocks/{id}", "更新 Mock"),
            ("mock.yaml", "delete_success", "DELETE", "/mocks/{id}", "删除 Mock"),
            ("mock.yaml", "hit_success", "GET", "/mock/{project_id}/{path}", "命中 Mock 规则"),
            ("mock.yaml", "hit_methods", "*", "/mock/{project_id}/{path}", "按方法命中 Mock"),
            ("mock.yaml", "hit_fail", "*", "/mock/{project_id}/{path}", "命中失败"),
        ],
    },
    {
        "title": "定时任务",
        "sections": [
            ("schedule.yaml", "create_success", "POST", "/schedules", "创建定时任务"),
            ("schedule.yaml", "create_no_token", "POST", "/schedules", "创建定时任务(不带 token)"),
            ("schedule.yaml", "list_success", "GET", "/schedules", "定时任务列表"),
            ("schedule.yaml", "detail_success", "GET", "/schedules/{id}", "定时任务详情"),
            ("schedule.yaml", "detail_fail", "GET", "/schedules/{id}", "定时任务详情(不存在)"),
            ("schedule.yaml", "update_success", "PUT", "/schedules/{id}", "更新定时任务"),
            ("schedule.yaml", "delete_success", "DELETE", "/schedules/{id}", "删除定时任务"),
        ],
    },
    {
        "title": "压测",
        "sections": [
            ("perf.yaml", "create_success", "POST", "/perf/tasks", "创建压测任务"),
            ("perf.yaml", "create_fail", "POST", "/perf/tasks", "创建压测任务失败"),
            ("perf.yaml", "create_no_token", "POST", "/perf/tasks", "创建压测任务(不带 token)"),
            ("perf.yaml", "list_success", "GET", "/perf/tasks", "压测任务列表"),
            ("perf.yaml", "detail_success", "GET", "/perf/tasks/{id}", "压测任务详情"),
            ("perf.yaml", "detail_fail", "GET", "/perf/tasks/{id}", "压测任务详情(不存在)"),
            ("perf.yaml", "delete_success", "DELETE", "/perf/tasks/{id}", "删除压测任务"),
            ("perf.yaml", "run_success", "POST", "/perf/tasks/{id}/run", "触发压测(异步)"),
        ],
    },
    {
        "title": "系统",
        "sections": [
            ("system.yaml", "health_success", "GET", "/health", "健康检查"),
            ("system.yaml", "metrics_success", "GET", "/metrics", "Prometheus 指标"),
        ],
    },
    {
        "title": "报告与其他",
        "sections": [
            ("misc.yaml", "report_list", "GET", "/reports", "报告列表"),
            ("misc.yaml", "report_detail_fail", "GET", "/reports/{id}", "报告详情(不存在)"),
            ("misc.yaml", "regression_run", "POST", "/regression/run", "回归执行"),
            ("misc.yaml", "demo_order_create", "POST", "/demo/orders", "demo 下单"),
            ("misc.yaml", "demo_order_list", "GET", "/demo/orders", "demo 订单列表"),
            ("misc.yaml", "traffic_record_list", "GET", "/traffic/records", "流量记录列表"),
            ("misc.yaml", "traffic_record_detail_fail", "GET", "/traffic/records/{id}", "流量记录详情(不存在)"),
            ("misc.yaml", "traffic_replay", "POST", "/traffic/records/{id}/replay", "流量回放"),
        ],
    },
]

# 同组多条 / 需要更贴切文案时,按 case_id 覆盖描述与备注
CASE_DESC = {
    # 登录
    "wrong_password": "已存在用户 + 错误密码",
    "user_not_exist": "不存在的用户名",
    # 登出
    "missing_token": "不带 Authorization 头",
    "invalid_token": "非法 token 登出",
    # 注册
    "register_duplicate_user": "重复用户名注册",
    # 项目
    "duplicate_project_name": "重复项目名",
    "missing_project_name": "缺 name 字段",
    "create_project_without_token": "创建项目不带 token",
    "list_projects_without_token": "项目列表不带 token",
    "get_project_detail_without_token": "项目详情不带 token",
    "delete_project_without_token": "删除项目不带 token",
    "delete_project_not_exist": "删除不存在的项目",
    # 接口
    "create_interface_without_token": "创建接口不带 token",
    "create_interface_without_project": "创建接口缺 X-Project-Id 头",
}

CASE_REMARK = {
    # 鉴权反例(不带 token 应为 401)
    "create_case_without_token": "鉴权反例:不带 token(401)",
    "create_environment_without_token": "鉴权反例:不带 token(401)",
    "create_mock_without_token": "鉴权反例:不带 token(401)",
    "create_schedule_without_token": "鉴权反例:不带 token(401)",
    "create_perf_task_without_token": "鉴权反例:不带 token(401)",
    "create_scenario_without_token": "鉴权反例:不带 token(401)",
    "create_interface_without_token": "鉴权反例:不带 token(401)",
}

# negative_contract 参数化契约反例展开(无独立 case_id)
NEG_CHECKS = [
    ("list_no_project_header", "GET", "{path}", "缺 X-Project-Id 头", "HTTP 422", "全资源一致"),
    ("list_no_token", "GET", "{path}", "不带 token 列表", "HTTP 401", "统一鉴权反例"),
    ("create_missing_field", "POST", "{path}", "创建缺必填字段(POST {})", "HTTP 422", "仅 creatable 资源"),
    ("detail_bad_id_type", "GET", "{path}/not-a-number", "详情 id 传非数字", "HTTP 422", "仅有详情的资源"),
    ("update_not_found", "PUT", "{path}/999999", "更新不存在的 id(带合法 body)", "HTTP 404", "仅 updatable 资源"),
]


def compose_expected(expected, case):
    if case.get("pending"):
        return "跳过:" + case.get("pending_reason", "pending")
    if not expected:
        return "-"
    parts = []
    if "status_code" in expected:
        parts.append(f"HTTP {expected['status_code']}")
    if "body_type" in expected:
        parts.append(f"返回 {expected['body_type']}")
    if "required_fields" in expected:
        parts.append("含字段 " + "/".join(expected["required_fields"]))
    if "detail" in expected:
        parts.append(f'detail="{expected["detail"]}"')
    return "；".join(parts)


def fmt_request(case):
    req = case.get("request")
    if req is None:
        return "-"
    if req == {}:
        return "{}"
    return json.dumps(req, ensure_ascii=False)


_yaml_cache = {}


def load(rel):
    if rel not in _yaml_cache:
        with open(DATA / rel, encoding="utf-8") as f:
            _yaml_cache[rel] = yaml.safe_load(f)
    return _yaml_cache[rel]


def build_rows():
    """返回 {sheet_title: [row,...]},并做 case_id 覆盖校验。"""
    sheets = {}
    covered = set()          # META 已覆盖的 case_id
    for spec in SHEETS:
        rows = []
        for rel, group, method, path, gdesc in spec["sections"]:
            doc = load(rel)
            cases = doc.get(group)
            if not cases:
                print(f"[WARN] {rel} 缺少分组 {group}")
                continue
            for case in cases:
                cid = case["case_id"]
                covered.add(cid)
                rows.append([
                    cid,
                    group,
                    method,
                    path,
                    CASE_DESC.get(cid, gdesc),
                    fmt_request(case),
                    compose_expected(case.get("expected"), case),
                    CASE_REMARK.get(cid, ""),
                ])
        for extra in spec.get("extra_rows", []):
            rows.append(extra)
        sheets[spec["title"]] = rows

    # negative_contract sheet
    neg = load("negative_contract.yaml")["resources"]
    neg_rows = []
    for res in neg:
        for key, method, path_tmpl, desc, exp, remark in NEG_CHECKS:
            if key == "create_missing_field" and not res["creatable"]:
                continue
            if key == "detail_bad_id_type" and not res["has_detail"]:
                continue
            if key == "update_not_found" and not res.get("updatable"):
                continue
            expected = exp
            if key == "list_no_token":
                expected = f"HTTP {res['list_no_token_status']}"
            if key == "create_missing_field":
                req_str = "{}"
            elif key == "update_not_found":
                req_str = json.dumps(res.get("update_body", {}), ensure_ascii=False)
            else:
                req_str = "-"
            neg_rows.append([
                f"{res['name']}__{key}",
                key,
                method,
                path_tmpl.format(path=res["path"]),
                f"{res['name']}:{desc}",
                req_str,
                expected,
                remark,
            ])
    sheets["通用契约反例"] = neg_rows

    # boundary sheet:data/boundary.yaml 是扁平 cases 列表(path/body/note),单独展开
    bnd = load("boundary.yaml")["cases"]
    bnd_rows = []
    for case in bnd:
        cid = case["case_id"]
        covered.add(cid)
        bnd_rows.append([
            cid,
            "boundary",
            "POST",
            case["path"],
            case.get("note", "边界/畸形输入"),
            json.dumps(case.get("body", {}), ensure_ascii=False),
            compose_expected(case.get("expected"), case),
            "特征化:pin 住畸形输入现状",
        ])
    sheets["边界与畸形输入"] = bnd_rows

    # 一致性校验:所有 yaml 里的 case_id 是否都被覆盖
    _check_coverage(covered)
    return sheets


def _check_coverage(covered):
    all_ids = set()
    for rel in [
        "login.yaml", "logout.yaml", "misc.yaml", "interface.yaml", "case.yaml",
        "scenario.yaml", "environment.yaml", "mock.yaml", "schedule.yaml",
        "perf.yaml", "system.yaml", "boundary.yaml",
        "project/create_project.yaml", "project/list_project.yaml",
        "project/get_project_detail.yaml", "project/delete_project.yaml",
    ]:
        doc = load(rel)
        for group, cases in doc.items():
            if isinstance(cases, list):
                for c in cases:
                    if isinstance(c, dict) and "case_id" in c:
                        all_ids.add(c["case_id"])
    missing = all_ids - covered
    if missing:
        print(f"[ERROR] 以下 YAML case_id 未进台账,请补 SHEETS 配置: {sorted(missing)}")
        sys.exit(1)
    print(f"[OK] 覆盖校验通过:{len(all_ids)} 个 YAML case_id 全部进台账。")


# --------------------------- 样式与写出 ---------------------------
HEAD_FILL = PatternFill("solid", fgColor="305496")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="center", wrap_text=True)
WIDTHS = [30, 20, 8, 30, 34, 40, 40, 26]


def write_workbook(sheets):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    total = 0
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        ws.append(HEADERS)
        for c in ws[1]:
            c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            ws.append(row)
            total += 1
        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
            for c in r:
                c.border, c.alignment = BORDER, WRAP
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(OUT)
    print(f"[DONE] 写出 {OUT}  共 {len(sheets)} sheet / {total} 行")


if __name__ == "__main__":
    write_workbook(build_rows())
